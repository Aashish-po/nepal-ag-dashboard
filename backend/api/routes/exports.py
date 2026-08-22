import csv
import re
import unicodedata
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Optional, Sequence, Tuple

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import Row, select
from sqlalchemy.orm import Session

from api.db import get_db
from api.models.db_models import Crops, Districts, Forecasts, Yields

router = APIRouter()


def _csv_response(content: str, filename: str) -> StreamingResponse:
    return StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _safe_filename_part(value: str) -> str:
    ascii_value = (
        unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    )
    return re.sub(r"[^A-Za-z0-9._-]+", "_", ascii_value).strip("_") or "unknown"


@router.get("/export/yields")
def export_yields(
    db: Session = Depends(get_db),
    district_id: Optional[int] = Query(None),
    crop_id: Optional[int] = None,
    year_start: int = Query(2014),
    year_end: int = Query(2024),
    format: str = Query("csv", description="Only 'csv' supported"),
):
    if format != "csv":
        raise HTTPException(status_code=400, detail="Only 'csv' format is supported")

    stmt = (
        select(
            Districts.name.label("district"),
            Crops.name.label("crop"),
            Yields.year.label("year"),
            Yields.yield_kg_ha.label("yield_kg_ha"),
            Yields.production_mt.label("production_mt"),
            Yields.area_harvested_ha.label("area_harvested_ha"),
            Yields.data_source.label("data_source"),
            Yields.data_quality.label("data_quality"),
        )
        .join(Districts, Yields.district_id == Districts.id)
        .join(Crops, Yields.crop_id == Crops.id)
        .where(Yields.year >= year_start)
        .where(Yields.year <= year_end)
    )
    if district_id is not None:
        stmt = stmt.where(Yields.district_id == district_id)
    if crop_id is not None:
        stmt = stmt.where(Yields.crop_id == crop_id)

    results = db.execute(stmt).all()

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "District",
            "Crop",
            "Year",
            "Yield (kg/ha)",
            "Production (MT)",
            "Area (ha)",
            "Data Source",
            "Quality",
        ]
    )
    for row in results:
        writer.writerow(
            [
                row.district,
                row.crop,
                row.year,
                round(float(row.yield_kg_ha), 2) if row.yield_kg_ha is not None else "",
                (
                    round(float(row.production_mt), 2)
                    if row.production_mt is not None
                    else ""
                ),
                (
                    round(float(row.area_harvested_ha), 2)
                    if row.area_harvested_ha is not None
                    else ""
                ),
                row.data_source or "",
                row.data_quality or "",
            ]
        )

    filename = f"yields_{datetime.now().strftime('%Y%m%d')}.csv"
    return _csv_response(output.getvalue(), filename)


@router.get("/export/forecasts")
def export_forecasts(
    db: Session = Depends(get_db),
    district_id: int = Query(..., description="Required: district ID"),
    crop_id: int = Query(..., description="Required: crop ID"),
    months_ahead: int = Query(
        12, ge=1, le=36, description="Forecast horizon (1-36 months)"
    ),
    format: str = Query("excel", description="Only 'excel' supported"),
):
    if format != "excel":
        raise HTTPException(
            status_code=400,
            detail="Only 'excel' format is supported",
        )

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    district = db.get(Districts, district_id)
    if not district:
        raise HTTPException(status_code=404, detail="District not found")

    crop = db.get(Crops, crop_id)
    if not crop:
        raise HTTPException(status_code=404, detail="Crop not found")

    # Historical data
    hist_stmt = (
        select(
            Yields.year,
            Yields.yield_kg_ha,
            Yields.production_mt,
        )
        .where(Yields.district_id == district_id)
        .where(Yields.crop_id == crop_id)
        .where(Yields.yield_kg_ha.isnot(None))
        .order_by(Yields.year)
    )
    hist_results = db.execute(hist_stmt).all()

    # Forecast data
    fc_stmt = (
        select(
            Forecasts.forecast_month,
            Forecasts.forecast_yield_kg_ha,
            Forecasts.lower_ci_95,
            Forecasts.upper_ci_95,
            Forecasts.forecast_model,
            Forecasts.rmse_kg_ha,
            Forecasts.mae_kg_ha,
            Forecasts.mape_pct,
        )
        .where(Forecasts.district_id == district_id)
        .where(Forecasts.crop_id == crop_id)
        .order_by(Forecasts.forecast_month)
    )

    fc_results: Sequence[
        Row[
            Tuple[
                date,
                Optional[float],
                Optional[float],
                Optional[float],
                Optional[str],
                Optional[float],
                Optional[float],
                Optional[float],
            ]
        ]
    ] = db.execute(fc_stmt).all()

    wb = Workbook()

    def _num(value, default=""):
        return round(float(value), 2) if value is not None else default

    # Sheet 1: Historical Data
    ws_hist = wb.active
    assert ws_hist is not None
    ws_hist.title = "Historical Data"
    ws_hist.append(["Year", "Yield (kg/ha)", "Production (MT)"])

    for row in hist_results:
        ws_hist.append(
            [
                row[0],
                _num(row[1]),
                _num(row[2]),
            ]
        )

    # Sheet 2: Forecasts
    ws_fc = wb.create_sheet("Forecasts")
    assert ws_fc is not None
    ws_fc.append(
        [
            "Forecast Month",
            "Forecast Yield (kg/ha)",
            "Lower 95% CI",
            "Upper 95% CI",
            "Model",
        ]
    )

    for hist_row in hist_results:
        ws_hist.append(
            [
                hist_row[0],
                _num(hist_row[1]),
                _num(hist_row[2]),
            ]
        )

    # Sheet 2: Forecasts
    ws_fc = wb.create_sheet("Forecasts")
    assert ws_fc is not None
    ws_fc.append(
        [
            "Forecast Month",
            "Forecast Yield (kg/ha)",
            "Lower 95% CI",
            "Upper 95% CI",
            "Model",
        ]
    )

    for forecast_row in fc_results[:months_ahead]:
        ws_fc.append(
            [
                forecast_row[0],
                _num(forecast_row[1]),
                _num(forecast_row[2]),
                _num(forecast_row[3]),
                forecast_row[4] or "",
            ]
        )
    # Sheet 3: Model Diagnostics
    ws_diag = wb.create_sheet("Model Diagnostics")
    assert ws_diag is not None

    if fc_results:
        ws_diag.append(["Metric", "Value"])
        ws_diag.append(["Model", fc_results[0][4] or "N/A"])
        ws_diag.append(["RMSE (kg/ha)", _num(fc_results[0][5], "N/A")])
        ws_diag.append(["MAE (kg/ha)", _num(fc_results[0][6], "N/A")])
        ws_diag.append(["MAPE (%)", _num(fc_results[0][7], "N/A")])

    # Sheet 4: Chart Data
    ws_chart = wb.create_sheet("Chart")
    assert ws_chart is not None
    ws_chart.append(
        [
            "Month",
            "Historical Yield",
            "Forecast",
            "Lower CI",
            "Upper CI",
        ]
    )

    # Write historical data to chart sheet
    for hist_row in hist_results:
        ws_chart.append(
            [
                hist_row[0],
                _num(hist_row[1]),
                None,  # No forecast value for historical data
                None,  # No lower CI for historical data
                None,  # No upper CI for historical data
            ]
        )

    # Write forecast data to chart sheet
    for forecast_row in fc_results[:months_ahead]:
        ws_chart.append(
            [
                forecast_row[0],
                None,  # No historical yield for forecast data
                _num(forecast_row[1]),
                _num(forecast_row[2]),
                _num(forecast_row[3]),
            ]
        )
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            col_idx = column[0].column
            if col_idx is not None:
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = (
        f"forecast_{_safe_filename_part(district.name)}"
        f"_{_safe_filename_part(crop.name)}"
        f"_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
